"""Excel (.xlsx) sheet converter (import + export).

Port of BakingSheet.Converters.Excel/ExcelSheetConverter.cs. Uses openpyxl.
Each workbook tab is a sheet (tab name may be ``SheetName.SubName`` for partial
sheets). Cells are stringified via the format provider, mirroring C#
``Convert.ToString``.
"""
from __future__ import annotations

import datetime
import os
from pathlib import Path
from typing import Any, List, Optional

from .._internal.config import parse_sheet_name
from ..raw.importer import RawSheetConverter, RawSheetExporterPage, RawSheetImporterPage


def _stringify_cell(value: Any, format_provider: Any = None) -> Optional[str]:
    """Coerce a native openpyxl cell value to a string, like C# Convert.ToString."""
    if value is None:
        return None
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime.datetime):
        # mirror C# DateTime cell rendering (local, no tz suffix)
        if value.microsecond:
            return value.strftime("%Y-%m-%d %H:%M:%S.%f")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.time):
        return value.strftime("%H:%M:%S")
    return str(value)


class _ExcelImporterPage(RawSheetImporterPage):
    def __init__(self, ws, sub_name: Optional[str], format_provider: Any) -> None:
        self._ws = ws
        self.sub_name = sub_name
        self._format_provider = format_provider
        # pre-read the used range into a 2D list for simple indexing
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([_stringify_cell(v, format_provider) for v in row])
        self._cells = rows

    def get_cell(self, col: int, row: int) -> Optional[str]:
        if row < 0 or row >= len(self._cells):
            return None
        r = self._cells[row]
        if col < 0 or col >= len(r):
            return None
        return r[col]


class ExcelSheetConverter(RawSheetConverter):
    """Import/export Excel workbooks. Port of ``ExcelSheetConverter``."""

    def __init__(
        self,
        load_path: str,
        timezone=None,
        format_provider=None,
        split_header: bool = False,
    ) -> None:
        super().__init__(timezone, format_provider, split_header)
        self._load_path = load_path
        self._pages: "dict[str, list[RawSheetImporterPage]]" = {}

    # -- import ---------------------------------------------------------
    def load_data(self) -> bool:
        import openpyxl

        path = Path(self._load_path)
        if path.is_dir():
            files = sorted(path.glob("*.xlsx"))
        elif path.is_file():
            files = [path]
        else:
            return True
        for file in files:
            if file.name.startswith("~$"):
                continue
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            for ws in wb.worksheets:
                if ws.title.startswith("$"):
                    continue
                sheet_name, sub_name = parse_sheet_name(ws.title)
                page = _ExcelImporterPage(ws, sub_name, self.format_provider)
                self._pages.setdefault(sheet_name, []).append(page)
            wb.close()
        return True

    def get_pages(self, sheet_name: str) -> List[RawSheetImporterPage]:
        return self._pages.get(sheet_name, [])

    # -- export ---------------------------------------------------------
    def create_page(self, sheet_name: str) -> RawSheetExporterPage:
        return RawSheetExporterPage(sheet_name)

    def save_data(self, pages: "list[tuple[str, RawSheetExporterPage]]") -> bool:  # type: ignore[override]
        import openpyxl

        os.makedirs(self._load_path, exist_ok=True)
        for name, page in pages:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = name
            for r, row in enumerate(page.to_grid()):
                for c, val in enumerate(row):
                    if val is not None and val != "":
                        ws.cell(row=r + 1, column=c + 1, value=val)
            wb.save(os.path.join(self._load_path, f"{name}.xlsx"))
            wb.close()
        return True
